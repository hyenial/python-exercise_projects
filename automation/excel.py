import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('python-spreadsheet.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1):
	cell = sheet.cell(row, 3)
	corrected_price = float(cell.value.replace('$','')) * 0.9
	corrected_price_cell = sheet.cell(row, 4)
	corrected_price_cell.value = corrected_price

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('python-spreadsheet2.xlsx')
